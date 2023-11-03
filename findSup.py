import openpyxl
from xml.dom import minidom
import xml.etree.ElementTree as ET

from datetime import datetime
start_time = datetime.now()

total_ibans_found = 0
total_suppliers_found = 0
total_invoices_found = 0
total_lines = 0

def main(input_workbook):
    workbook = openpyxl.load_workbook(f'Unreconciled Database_2023-11-02.xlsx')
    worksheet = workbook.active

    anicura_clinics = openpyxl.load_workbook('anicura_ibans.xlsx')
    anicura_clinics_worksheet = anicura_clinics.active

    suppliers_xml = minidom.parse('suppliers_data.xml')
    suppliers = suppliers_xml.getElementsByTagName('record')

    global total_lines
    results = []

    for row_index, row in enumerate(worksheet.iter_rows(min_row=2, values_only=True)):
        bank_account = row[0]
        date = row[9]
        status = row[22]
        ledger = row[13]

        if bank_account is None:
            break

        if date == None or date.year != 2023 or status not in ['Not Started', 'Stuck'] or any(code in ledger for code in ['DK', 'SE']):
            continue

        amount = row[5]
        description = row[11]

        if description == None:
            continue

        clinic = row[12]
        id = row[14]

        row_to_fill = row_index + 2
        worksheet.cell(row=row_to_fill, column=30, value=None)
        worksheet.cell(row=row_to_fill, column=31, value=None) 
        
        found_iban = find_iban(description, anicura_clinics_worksheet)
        found_supplier = find_supplier(suppliers, found_iban)

        line = {
            'ledger': ledger, 
            'clinic': clinic, 
            'id': id, 
            'description': description, 
            'amount': amount, 
            'iban': found_iban, 
            'supplier': found_supplier
        }
        
        results.append(line)
        total_lines += 1

        worksheet.cell(row=row_to_fill, column=30, value=found_supplier)
        worksheet.cell(row=row_to_fill, column=31, value=found_iban)      

    worksheet.views.sheetView[0].pane = openpyxl.worksheet.views.Pane()
    workbook.save(filename=f'{input_workbook}_filled.xlsx')

    write_list_to_xml(results, f'{input_workbook}_matching_results.xml')

def find_iban(description, clinic_ibans):
    global total_ibans_found
    ibans = []
    
    iban_len = {
        'NO':15,
        'BE':16,
        'NL':18, 
        'DK':18, 
        'SI':19, 
        'AT':20, 
        'CH':21, 
        'DE':22, 
        'GB':22, 
        'ES':24, 
        'SE':24, 
        'SK':24, 
        'PT':25, 
        'FR':27, 
        'IT':27, 
        'SM':27, 
        'PL':28,
        'LU':20,
        'IE':22
    }

    for code in iban_len:
        code_positions = []
        start = 0

        while start < len(description):
            position = description.find(code, start)

            if position == -1:
                break

            code_positions.append(position)
            start = position + len(code)

        for position in code_positions:
            next_two_signs = description[(position+2):(position+4)]

            if next_two_signs.isdigit():
                iban = description[position:position + iban_len[code]]

                if ' ' not in iban and '.' not in iban:
                    ibans.append(iban)
            else:
                continue

    for index, iban in enumerate(ibans):

        if iban != None:
            for row in clinic_ibans.iter_rows(min_row=2, values_only=True):
                anicura_iban = row[4]

                if anicura_iban == iban:
                    del ibans[index]
                    break
            
            if len(ibans) != 0:
                total_ibans_found += 1
                return ibans[0]
            else:
                return "None"

def find_supplier (suppliers, iban):
    global total_suppliers_found

    for supplier in suppliers:
        supplier_iban = supplier.getElementsByTagName('IBAN')[0].childNodes[0].nodeValue
        
        if supplier_iban == iban:
            supplier_name = supplier.getElementsByTagName('SUPPLIER_NAME')[0].childNodes[0].nodeValue
            total_suppliers_found += 1
            break
        else:
            supplier_name = None

    return supplier_name

def write_list_to_xml(list, file_name):
    root = ET.Element('data')

    for item in list:
        record = ET.SubElement(root, 'record')

        ledger = ET.SubElement(record, 'LEDGER')
        ledger.text = str(item['ledger'])

        clinic = ET.SubElement(record, 'CLINIC')
        clinic.text = str(item['clinic'])

        id = ET.SubElement(record, "ID")
        id.text = str(item['id'])

        description = ET.SubElement(record, 'DESCRIPTION')
        description.text = str(item['description'])

        amount = ET.SubElement(record, 'AMOUNT')
        amount.text = str(item['amount'])

        iban = ET.SubElement(record, 'IBAN')
        iban.text = str(item['iban'])

        supplier = ET.SubElement(record, 'SUPPLIER')
        supplier.text = str(item['supplier'])
    
    tree = ET.ElementTree(root)

    xml_file = file_name
    tree.write(xml_file)

main('Workbook_13')

end_time = datetime.now()
print(f'lines: {total_lines} / ibans found: {total_ibans_found} / suppliers found: {total_suppliers_found}')
print(f'Ratio of ibans found: {round(total_ibans_found/total_lines, 2)}')
print(f'Ratio of suppliers found: {round(total_suppliers_found/total_lines, 2)}')
print(f'Ratio of suppliers/ibans found: {round(total_suppliers_found/total_ibans_found, 2)}')
print('Duration: {}'.format(end_time - start_time))