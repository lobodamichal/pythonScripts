'''
- install Visual Studio:
https://code.visualstudio.com/?wt.mc_id=vscom_downloads
- install python (v. 3.11):
https://www.python.org/downloads/
    or (in CMD)
> python
    then download and install
- install openpyxl package (CMD):
> pip install openpyxl

How to run script:
- go to CMD
- navigate to directory with script file using (basic CMD tutorial: https://riptutorial.com/cmd/example/8646/navigating-in-cmd): 
> cd <directory name>
- if you are in correct directory run command:
> python <filename>
'''

import openpyxl
from xml.dom import minidom
import re

#STATS
from datetime import datetime
start_time = datetime.now()
total_ibans_found = 1
total_suppliers_found = 0
total_invoices_found = 0
total_lines = 1

def main():

    #----#
    #LOAD EXCEL SHEETS AND XML'S
    #<file name> is in same directory as script or set the path with name
    #this will change arguments to be hard coded here
    #if not it's not then inputs will appear in 

    input_workbook = None
    input_suppliers = None
    input_save_to = None
    input_parse = None

    while True:
        input_parse = input('do you want to enter the names of files? <Y/N>: ')
        if (input_parse in ['Y', 'N', 'y', 'n']):
            break

    if input_parse in ['n', 'N']:
        #input_workbook = 'path/to/Workbook_X.xlsx'
        input_workbook = './Workbook_7.xlsx'
        #input_suppliers = 'path/to/Suppliers.xml'
        input_suppliers = './suppliers_data.xml'
        #input_save_to = 'path/to/SaveTo.xlsx'
        input_save_to = './filled/Workbook_7.xlsx'
    else:
        input_workbook = input('enter workbook path <path/to/Workbook_X.xlsx>: ')
        input_suppliers = input('enter suppliers path <path/to/Suppliers.xml>: ')
        input_save_to = input('enter where to save file <path/to/SaveTo.xlsx>: ')

    #WRITE ERROR HANDLING--------------->

    workbook = openpyxl.load_workbook(input_workbook)
    worksheet = workbook.active

    suppliers_xml = minidom.parse(input_suppliers)
    suppliers = suppliers_xml.getElementsByTagName('record')
    #----#

    #STATS
    global total_lines

    for row_index, row in enumerate(worksheet.iter_rows(min_row=2, values_only=True)):
        bank_account = row[0]
        #COLUMN 'A'

        if bank_account is None:
            break

        description = row[11]
        #COLUMN 'L'

        if description == None:
            continue

        row_to_fill = row_index + 2

        found_iban = find_iban(bank_account, description)
        found_supplier = find_supplier(suppliers, found_iban)

        #STATS
        total_lines += 1

        worksheet.cell(row=row_to_fill, column=30, value=found_supplier)
        #SUPPLIER IN COLUMN 'AD'
 
        worksheet.cell(row=row_to_fill, column=31, value=found_iban)      
        #IBAN IN COLUMN 'AE'

    worksheet.views.sheetView[0].pane = openpyxl.worksheet.views.Pane()
    workbook.save(filename=input_save_to)

def find_iban(bank_account, description):
    global total_ibans_found
    ibans = []
    
    iban_len = {
        'NO':15,
        'BE':16,
        'NL':18, 
        'DK':18, 
        'SI':19, 
        'AT':20, 
        'CH':21, 
        'DE':22, 
        'GB':22, 
        'ES':24, 
        'SE':24, 
        'SK':24, 
        'PT':25, 
        'FR':27, 
        'IT':27, 
        'SM':27, 
        'PL':28,
        'LU':20,
        'IE':22
    }

    for code in iban_len:
        code_positions = []
        start = 0

        while start < len(description):
            position = description.find(code, start)

            if position == -1:
                break

            code_positions.append(position)
            start = position + len(code)

        for position in code_positions:
            next_two_signs = description[(position+2):(position+4)]

            if next_two_signs.isdigit():
                iban = description[position:position + iban_len[code]]
                signs_pattern = re.compile(r'\W')
                check = re.search(signs_pattern, iban)

                if check is None:
                    ibans.append(iban)
            else:
                continue

    for index, iban in enumerate(ibans):
        if bank_account[-4:] == iban[-4:]:
            del ibans[index]
            break
    
        if len(ibans) != 0:

            #STATS
            total_ibans_found += 1

            return ibans[0]
        else:
            return "None"

def find_supplier (suppliers, iban):
    global total_suppliers_found

    for supplier in suppliers:
        supplier_iban = supplier.getElementsByTagName('IBAN')[0].childNodes[0].nodeValue
        
        if supplier_iban == iban:
            supplier_name = supplier.getElementsByTagName('SUPPLIER_NAME')[0].childNodes[0].nodeValue
            
            #STATS
            total_suppliers_found += 1

            break
        else:
            supplier_name = None

    return supplier_name

if __name__ == '__main__':
    main()
    
#LOG STATS
end_time = datetime.now()
print(f'lines: {total_lines} | ibans found: {total_ibans_found} | suppliers found: {total_suppliers_found}')
print(f'Ratio of ibans found: {round(total_ibans_found/total_lines, 2)}')
print(f'Ratio of suppliers found: {round(total_suppliers_found/total_lines, 2)}')
print(f'Ratio of suppliers/ibans found: {round(total_suppliers_found/total_ibans_found, 2)}')
print('Duration: {}'.format(end_time - start_time))